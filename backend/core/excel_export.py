from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from database.models import Job

HEADERS = (
    "Title",
    "Company",
    "Location",
    "Salary",
    "Posted At (UTC)",
    "URL",
    "Job ID",
    "Search Location",
    "Industry",
    "Portal",
)

_WIDTHS = (36, 28, 28, 18, 20, 48, 16, 16, 12, 12)
_BAD_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]+')


def _sheet_name(raw: str, used: set[str]) -> str:
    """Excel sheet title: max 31 chars, no illegal characters, unique."""
    base = _BAD_SHEET_CHARS.sub("-", (raw or "Unknown").strip()) or "Unknown"
    base = base[:31]
    name = base
    n = 2
    while name.lower() in used:
        suffix = f"_{n}"
        name = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(name.lower())
    return name


def _job_row(job: Job) -> list[str]:
    posted = job.posted_at
    if isinstance(posted, datetime):
        posted_str = posted.strftime("%Y-%m-%d %H:%M:%S")
    else:
        posted_str = str(posted or "")
    return [
        job.title or "",
        job.company_name or "",
        job.location or "",
        job.salary or "",
        posted_str,
        job.url or "",
        job.job_id or "",
        job.search_location or "",
        job.industry or "",
        job.source_portal or "",
    ]


def _style_sheet(ws) -> None:
    ws.append(list(HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    for idx, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def jobs_to_xlsx(jobs: Iterable[Job]) -> bytes:
    """One workbook; one sheet per search_location (crawl city)."""
    grouped: dict[str, list[Job]] = defaultdict(list)
    for job in jobs:
        key = (job.search_location or job.location or "Unknown").strip() or "Unknown"
        grouped[key].append(job)

    wb = Workbook()
    # Remove default sheet; create named sheets
    default = wb.active
    wb.remove(default)

    used_names: set[str] = set()
    if not grouped:
        ws = wb.create_sheet(_sheet_name("Empty", used_names))
        _style_sheet(ws)
    else:
        for loc_key in sorted(grouped.keys(), key=lambda s: s.lower()):
            ws = wb.create_sheet(_sheet_name(loc_key, used_names))
            _style_sheet(ws)
            for job in grouped[loc_key]:
                ws.append(_job_row(job))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
